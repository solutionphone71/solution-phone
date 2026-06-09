from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

red = "C0392B"; black = "000000"; light_gray = "F2F2F2"; mid_gray = "D9D9D9"

bold_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
bold_black = Font(name='Arial', size=10, bold=True, color='000000')
bold_red = Font(name='Arial', size=14, bold=True, color=red)
normal = Font(name='Arial', size=10)
italic = Font(name='Arial', size=10, italic=True, color='888888')

red_fill = PatternFill('solid', start_color=red, end_color=red)
black_fill = PatternFill('solid', start_color=black, end_color=black)
light_fill = PatternFill('solid', start_color=light_gray, end_color=light_gray)
mid_fill = PatternFill('solid', start_color=mid_gray, end_color=mid_gray)

# ─── ONGLET 1 : ACCESSOIRES & PIÈCES ─────────────────────────────────
ws1 = wb.active
ws1.title = "Accessoires-Pièces"
ws1.column_dimensions['A'].width = 55
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 16

ws1['A1'] = "SOLUTION PHONE — Inventaire au 31/03/2026"
ws1['A1'].font = bold_red
ws1.merge_cells('A1:D1')
ws1['A2'] = "1. Accessoires & pièces détachées"
ws1['A2'].font = Font(name='Arial', size=12, bold=True)
ws1.merge_cells('A2:D2')
ws1.row_dimensions[1].height = 24
ws1.row_dimensions[2].height = 20

headers = ['Désignation', 'Quantité', 'PU HT (€)', 'Montant HT (€)']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    c.font = bold_white; c.fill = black_fill
    c.alignment = Alignment(horizontal='center')

data = [
    ('Écrans iPhone', [
        ('Écran iPhone 15 Pro Max — qualité OEM Original', 5, 320.00),
        ('Écran iPhone 14 Pro Max — qualité OEM', 8, 280.00),
        ('Écran iPhone 14 Pro — qualité OEM', 10, 220.00),
        ('Écran iPhone 14 — qualité LPTS Prime', 8, 120.00),
        ('Écran iPhone 13 — qualité LPTS Prime', 12, 95.00),
        ('Écran iPhone 12 — qualité LPTS', 15, 65.00),
        ('Écran iPhone 11 — qualité HD ECO+', 10, 45.00),
        ('Écran iPhone XR — qualité LPTS', 8, 55.00),
        ('Écran iPhone XS Max — qualité OEM', 6, 85.00),
        ('Écran iPhone SE 2/3 — qualité LPTS', 12, 32.00),
    ]),
    ('Écrans Samsung', [
        ('Écran Samsung Galaxy S23 — Service Pack', 4, 220.00),
        ('Écran Samsung Galaxy S22 — Service Pack', 4, 180.00),
        ('Écran Samsung Galaxy A53 / A54', 6, 95.00),
        ('Écran Samsung Galaxy A14 / A24', 8, 55.00),
    ]),
    ('Batteries', [
        ('Batterie iPhone 14', 25, 22.00),
        ('Batterie iPhone 13', 20, 18.00),
        ('Batterie iPhone 12', 25, 16.00),
        ('Batterie iPhone 11', 30, 14.00),
        ('Batterie iPhone XR', 20, 13.00),
        ('Batterie iPhone X / XS', 15, 12.00),
        ('Batterie iPhone SE 2/3', 12, 10.00),
        ('Batterie Samsung Galaxy S22 / S23', 10, 25.00),
        ('Batterie Samsung Galaxy S21', 8, 22.00),
    ]),
    ('Pièces détachées diverses', [
        ('Connecteurs de charge (Lightning + USB-C, tous modèles)', 30, 8.00),
        ('Vitres arrière iPhone 12-14', 20, 15.00),
        ('Nappes de charge', 15, 12.00),
        ('Caméras avant iPhone', 20, 20.00),
        ('Caméras arrière iPhone', 15, 28.00),
        ('Boutons home iPhone', 50, 4.00),
        ('Lecteurs SIM', 40, 5.00),
        ('Nappes power / volume', 30, 8.00),
        ("Joints d'étanchéité, visserie, petites pièces", 1, 320.00),
    ]),
    ('Coques, films & protections', [
        ('Coques silicone tous modèles', 100, 4.00),
        ('Coques renforcées anti-choc', 80, 5.00),
        ('Verres trempés tous modèles', 60, 3.00),
        ('Films hydrogel anti-casse', 100, 2.00),
        ('Coques officielles Apple / Samsung', 40, 12.00),
        ('Coques fashion / designer', 30, 11.00),
    ]),
    ('Accessoires', [
        ('Chargeurs USB-C 20W', 30, 12.00),
        ('Chargeurs Lightning', 25, 9.00),
        ('Câbles USB-C / Lightning', 50, 5.00),
        ('Écouteurs filaires', 20, 6.00),
        ('Écouteurs Bluetooth', 15, 25.00),
        ('Supports voiture / muraux', 20, 8.00),
    ]),
]

row = 5
subtotal_rows = []
for category, items in data:
    c = ws1.cell(row=row, column=1, value="▸ " + category)
    c.font = bold_black; c.fill = mid_fill
    for col in range(1, 5):
        ws1.cell(row=row, column=col).fill = mid_fill
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    start = row
    for desig, qty, pu in items:
        ws1.cell(row=row, column=1, value=desig).font = normal
        ws1.cell(row=row, column=2, value=qty).font = normal
        ws1.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws1.cell(row=row, column=3, value=pu).font = normal
        ws1.cell(row=row, column=3).number_format = '#,##0.00 €'
        ws1.cell(row=row, column=4, value="=B" + str(row) + "*C" + str(row)).font = normal
        ws1.cell(row=row, column=4).number_format = '#,##0.00 €'
        row += 1
    ws1.cell(row=row, column=1, value="Sous-total " + category).font = bold_black
    for col in range(1, 5):
        ws1.cell(row=row, column=col).fill = light_fill
    ws1.cell(row=row, column=4, value="=SUM(D" + str(start) + ":D" + str(row-1) + ")").font = bold_black
    ws1.cell(row=row, column=4).number_format = '#,##0.00 €'
    subtotal_rows.append(row)
    row += 2

row += 1
ws1.cell(row=row, column=1, value="TOTAL Section 1 — Accessoires & pièces").font = bold_white
for col in range(1, 5):
    ws1.cell(row=row, column=col).fill = red_fill
sum1 = "+".join(["D" + str(r) for r in subtotal_rows])
ws1.cell(row=row, column=4, value="=" + sum1).font = bold_white
ws1.cell(row=row, column=4).number_format = '#,##0.00 €'
ws1.row_dimensions[row].height = 22
total1_row = row

# ─── ONGLET 2 : SAMSUNG MOTOROLA NEUFS ───────────────────────────────
ws2 = wb.create_sheet("Samsung-Motorola neufs")
ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 16

ws2['A1'] = "SOLUTION PHONE — Inventaire au 31/03/2026"
ws2['A1'].font = bold_red
ws2.merge_cells('A1:D1')
ws2['A2'] = "2. Smartphones neufs Samsung & Motorola"
ws2['A2'].font = Font(name='Arial', size=12, bold=True)
ws2.merge_cells('A2:D2')
ws2.row_dimensions[1].height = 24
ws2.row_dimensions[2].height = 20

for col, h in enumerate(headers, 1):
    c = ws2.cell(row=4, column=col, value=h)
    c.font = bold_white; c.fill = black_fill
    c.alignment = Alignment(horizontal='center')

neufs = [
    ('Samsung', [
        ('Samsung Galaxy S24 — 128 Go — Noir', 1, 690.00),
        ('Samsung Galaxy A55 5G — 256 Go — Bleu nuit', 1, 380.00),
        ('Samsung Galaxy A35 5G — 128 Go — Noir / Bleu', 2, 240.00),
        ('Samsung Galaxy A15 — 128 Go — Noir', 2, 150.00),
        ('Samsung Galaxy A05 — 64 Go — Noir / Argent', 2, 95.00),
    ]),
    ('Motorola', [
        ('Motorola Razr 40 — 256 Go — Vert sauge', 1, 540.00),
        ('Motorola Edge 50 Fusion — 256 Go', 1, 320.00),
        ('Motorola Moto G84 — 256 Go', 1, 235.00),
        ('Motorola Moto G54 5G — 256 Go', 1, 175.00),
        ('Motorola Moto G34 — 128 Go', 2, 130.00),
    ]),
]

row = 5
subtotal2 = []
for cat, items in neufs:
    c = ws2.cell(row=row, column=1, value="▸ " + cat)
    c.font = bold_black; c.fill = mid_fill
    for col in range(1, 5):
        ws2.cell(row=row, column=col).fill = mid_fill
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    start = row
    for d, q, p in items:
        ws2.cell(row=row, column=1, value=d).font = normal
        ws2.cell(row=row, column=2, value=q).font = normal
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=3, value=p).font = normal
        ws2.cell(row=row, column=3).number_format = '#,##0.00 €'
        ws2.cell(row=row, column=4, value="=B" + str(row) + "*C" + str(row)).font = normal
        ws2.cell(row=row, column=4).number_format = '#,##0.00 €'
        row += 1
    ws2.cell(row=row, column=1, value="Sous-total " + cat).font = bold_black
    for col in range(1, 5):
        ws2.cell(row=row, column=col).fill = light_fill
    ws2.cell(row=row, column=4, value="=SUM(D" + str(start) + ":D" + str(row-1) + ")").font = bold_black
    ws2.cell(row=row, column=4).number_format = '#,##0.00 €'
    subtotal2.append(row)
    row += 2

row += 1
ws2.cell(row=row, column=1, value="TOTAL Section 2 — Samsung & Motorola neufs").font = bold_white
for col in range(1, 5):
    ws2.cell(row=row, column=col).fill = red_fill
sum2 = "+".join(["D" + str(r) for r in subtotal2])
ws2.cell(row=row, column=4, value="=" + sum2).font = bold_white
ws2.cell(row=row, column=4).number_format = '#,##0.00 €'
ws2.row_dimensions[row].height = 22
total2_row = row

# ─── ONGLET 3 : SMARTPHONES OCCASION ─────────────────────────────────
ws3 = wb.create_sheet("Smartphones occasion")
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 22
ws3.column_dimensions['F'].width = 25
ws3.column_dimensions['G'].width = 13
ws3.column_dimensions['H'].width = 14

ws3['A1'] = "SOLUTION PHONE — Inventaire au 31/03/2026"
ws3['A1'].font = bold_red
ws3.merge_cells('A1:H1')
ws3['A2'] = "3. Smartphones d'occasion en stock au 31/03/2026"
ws3['A2'].font = Font(name='Arial', size=12, bold=True)
ws3.merge_cells('A2:H2')
ws3['A3'] = "→ Exécute sql/inventaire-31-03-2026.sql dans Supabase → exporte CSV → colle les lignes ici (à partir de la ligne 6)"
ws3['A3'].font = italic
ws3.merge_cells('A3:H3')
ws3.row_dimensions[1].height = 24
ws3.row_dimensions[2].height = 20

occ_headers = ['Modèle', 'Stockage Go', 'Grade', 'Couleur', 'IMEI', 'Vendeur / Fournisseur', 'Date achat', 'Prix achat HT (€)']
for col, h in enumerate(occ_headers, 1):
    c = ws3.cell(row=5, column=col, value=h)
    c.font = bold_white; c.fill = black_fill
    c.alignment = Alignment(horizontal='center')

# 50 lignes vides pour remplir
for r in range(6, 56):
    ws3.cell(row=r, column=8).number_format = '#,##0.00 €'

ws3.cell(row=57, column=1, value="TOTAL Section 3 — Smartphones occasion").font = bold_white
for col in range(1, 9):
    ws3.cell(row=57, column=col).fill = red_fill
ws3.cell(row=57, column=8, value="=SUM(H6:H56)").font = bold_white
ws3.cell(row=57, column=8).number_format = '#,##0.00 €'
ws3.row_dimensions[57].height = 22

# ─── ONGLET 0 : RÉCAPITULATIF GÉNÉRAL ────────────────────────────────
ws4 = wb.create_sheet("Récap général", 0)
ws4.column_dimensions['A'].width = 55
ws4.column_dimensions['B'].width = 22

ws4['A1'] = "SOLUTION PHONE"
ws4['A1'].font = Font(name='Arial', size=18, bold=True, color=red)
ws4.merge_cells('A1:B1')
ws4['A2'] = "Inventaire valorisé au 31 mars 2026"
ws4['A2'].font = Font(name='Arial', size=14, bold=True)
ws4.merge_cells('A2:B2')
ws4['A3'] = "EURL Solution Phone — 21 rue Gambetta — 71000 Mâcon"
ws4['A3'].font = italic
ws4.merge_cells('A3:B3')
ws4['A4'] = "Mode de valorisation : coût d'acquisition HT (Art. R123-179)"
ws4['A4'].font = italic
ws4.merge_cells('A4:B4')
ws4.row_dimensions[1].height = 30
ws4.row_dimensions[2].height = 22

ws4['A6'] = "Catégorie"
ws4['B6'] = "Montant HT (€)"
ws4['A6'].font = bold_white; ws4['B6'].font = bold_white
ws4['A6'].fill = black_fill; ws4['B6'].fill = black_fill
ws4['B6'].alignment = Alignment(horizontal='right')

ws4['A7'] = "1. Accessoires & pièces détachées"
ws4['B7'] = "='Accessoires-Pièces'!D" + str(total1_row)
ws4['B7'].number_format = '#,##0.00 €'
ws4['B7'].alignment = Alignment(horizontal='right')

ws4['A8'] = "2. Smartphones neufs Samsung & Motorola"
ws4['B8'] = "='Samsung-Motorola neufs'!D" + str(total2_row)
ws4['B8'].number_format = '#,##0.00 €'
ws4['B8'].alignment = Alignment(horizontal='right')

ws4['A9'] = "3. Smartphones d'occasion (depuis Supabase)"
ws4['B9'] = "='Smartphones occasion'!H57"
ws4['B9'].number_format = '#,##0.00 €'
ws4['B9'].alignment = Alignment(horizontal='right')

ws4['A11'] = "TOTAL INVENTAIRE HT"
ws4['A11'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ws4['A11'].fill = red_fill
ws4['B11'] = "=B7+B8+B9"
ws4['B11'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ws4['B11'].fill = red_fill
ws4['B11'].number_format = '#,##0.00 €'
ws4['B11'].alignment = Alignment(horizontal='right')
ws4.row_dimensions[11].height = 28

ws4['A14'] = "Établi par : Sébastien Cannard — Gérant"
ws4['A14'].font = bold_black
ws4['A15'] = "Signature :  ____________________________________"
ws4['A17'] = "Date de l'inventaire physique : 31 mars 2026"
ws4['A18'] = "Lieu : 21 rue Gambetta, 71000 Mâcon"
ws4['A20'] = "Vérifié par : __________________________________"

# Frozen panes pour faciliter le scroll
ws1.freeze_panes = 'A5'
ws2.freeze_panes = 'A5'
ws3.freeze_panes = 'A6'

wb.save('inventaire-31-03-2026.xlsx')
print("FICHIER CRÉÉ : inventaire-31-03-2026.xlsx")
print(f"Total Section 1 row: {total1_row}")
print(f"Total Section 2 row: {total2_row}")
print(f"Sous-totaux Section 1 lignes : {subtotal_rows}")
print(f"Sous-totaux Section 2 lignes : {subtotal2}")
