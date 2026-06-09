from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook('inventaire-31-03-2026.xlsx')

# Supprime entièrement l'ancien onglet et recrée propre
if 'Smartphones occasion' in wb.sheetnames:
    del wb['Smartphones occasion']
ws3 = wb.create_sheet('Smartphones occasion')

red = "C0392B"; black = "000000"; light_gray = "F2F2F2"
bold_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
bold_black = Font(name='Arial', size=10, bold=True, color='000000')
bold_red = Font(name='Arial', size=14, bold=True, color=red)
normal = Font(name='Arial', size=10)
italic = Font(name='Arial', size=10, italic=True, color='888888')
red_fill = PatternFill('solid', start_color=red, end_color=red)
black_fill = PatternFill('solid', start_color=black, end_color=black)
light_fill = PatternFill('solid', start_color=light_gray, end_color=light_gray)

ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 22

ws3['A1'] = "SOLUTION PHONE — Inventaire au 31/03/2026"
ws3['A1'].font = bold_red
ws3.merge_cells('A1:D1')
ws3['A2'] = "3. Smartphones d'occasion en stock au 31/03/2026 (vue agrégée par modèle)"
ws3['A2'].font = Font(name='Arial', size=12, bold=True)
ws3.merge_cells('A2:D2')
ws3['A3'] = "Source : extraction Supabase phones · 17 modèles · 19 smartphones"
ws3['A3'].font = italic
ws3.merge_cells('A3:D3')
ws3.row_dimensions[1].height = 24
ws3.row_dimensions[2].height = 20

headers = ['Modèle', 'Quantité', 'Prix moyen HT (€)', 'Valeur totale HT (€)']
for col, h in enumerate(headers, 1):
    c = ws3.cell(row=5, column=col, value=h)
    c.font = bold_white
    c.fill = black_fill
    c.alignment = Alignment(horizontal='center')

data = [
    ('IPHONE 14 PRO MAX',   2,  385.00),
    ('IPHONE 15 PRO MAX',   1,  600.00),
    ('Iphone 15',           2,  290.00),
    ('Samsung S24',         2,  265.00),
    ('IPHONE 13 PRO MAX',   1,  355.00),
    ('IPHONE 13',           1,  335.00),
    ('S23 Ultra',           1,  300.00),
    ('IPHONE 14 pro',       1,  300.00),
    ('iPhone 13 pro',       1,  200.00),
    ('SAMSUNG S22 ultra',   1,  169.98),
    ('iPhone 12 PO MAX',    1,  160.00),
    ('SAMSUNG S21',         1,  155.00),
    ('HONOR MAGIC 8 LITE',  1,  150.00),
    ('iPhone 14 Pro',       1,  130.00),
    ('iPhone 12',           1,  130.00),
    ('iphone 11',           1,   80.00),
    ('IPHONE 8 PLUS',       1,   50.00),
]

row = 6
for modele, qty, prix_moy in data:
    ws3.cell(row=row, column=1, value=modele).font = normal
    ws3.cell(row=row, column=2, value=qty).font = normal
    ws3.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=3, value=prix_moy).font = normal
    ws3.cell(row=row, column=3).number_format = '#,##0.00 €'
    ws3.cell(row=row, column=3).alignment = Alignment(horizontal='right')
    ws3.cell(row=row, column=4, value="=B" + str(row) + "*C" + str(row)).font = normal
    ws3.cell(row=row, column=4).number_format = '#,##0.00 €'
    ws3.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    if row % 2 == 0:
        for col in range(1, 5):
            ws3.cell(row=row, column=col).fill = light_fill
    row += 1

total_row = row + 1
ws3.cell(row=total_row, column=1, value="TOTAL Section 3 — Smartphones occasion").font = bold_white
for col in range(1, 5):
    ws3.cell(row=total_row, column=col).fill = red_fill
ws3.cell(row=total_row, column=2, value="=SUM(B6:B" + str(row-1) + ")").font = bold_white
ws3.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
ws3.cell(row=total_row, column=4, value="=SUM(D6:D" + str(row-1) + ")").font = bold_white
ws3.cell(row=total_row, column=4).number_format = '#,##0.00 €'
ws3.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
ws3.row_dimensions[total_row].height = 22

ws_recap = wb['Récap général']
ws_recap['B9'] = "='Smartphones occasion'!D" + str(total_row)
ws_recap['B9'].number_format = '#,##0.00 €'
ws_recap['B9'].alignment = Alignment(horizontal='right')

ws3.freeze_panes = 'A6'

wb.save('inventaire-31-03-2026.xlsx')
print(f"OK — Total Section 3 row: {total_row}")
